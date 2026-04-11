"""
Excel → Bronze loader.

Принцип: максимально сырые данные.
Каждая строка каждого листа → одна запись в bronze.excel_tables.
Никакой бизнес-логики: заголовки, пустые строки, merged cells — всё хранится as-is.

Метаданные на каждую строку:
  load_id, loaded_at, source_file, year, sheet_name, row_num,
  range_address, merged_cells_meta

Данные: col_0, col_1, ... col_N  (schema evolution при появлении новых колонок)

Производительность:
  - Каждый файл открывается один раз для данных и один раз для merged cells.
  - Все листы одного файла пишутся одним батчем (один Iceberg append на файл).
  - Уже загруженные листы определяются одним scan в начале, а не per-sheet.
"""

import json
import logging
import os
import re
import uuid
from datetime import datetime, timezone

import pyarrow as pa
from pyiceberg.catalog.sql import SqlCatalog
from pyiceberg.types import StringType

log = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data", "Дошколка")
CATALOG_DIR = os.path.join(BASE_DIR, "catalog")


# ── Каталог ────────────────────────────────────────────────────────────────────

def get_catalog() -> SqlCatalog:
    return SqlCatalog(
        "datalake",
        **{
            "uri": f"sqlite:///{CATALOG_DIR}/catalog.db",
            "warehouse": f"file://{CATALOG_DIR}/warehouse",
        },
    )


# ── Обнаружение файлов ─────────────────────────────────────────────────────────

def discover_files(data_dir: str) -> list[tuple[int, str]]:
    """Файлы в подпапках с именем = год: data_dir/2018/file.xlsx"""
    result = []
    for entry in sorted(os.scandir(data_dir), key=lambda e: e.name):
        if not entry.is_dir():
            continue
        try:
            year = int(entry.name)
        except ValueError:
            continue
        for fname in os.listdir(entry.path):
            if fname.lower().endswith((".xls", ".xlsx")) and not fname.startswith("~"):
                result.append((year, os.path.join(entry.path, fname)))
    return result


def discover_files_flat(data_dir: str) -> list[tuple[int, str]]:
    """Файлы прямо в папке, год из имени: data_dir/Prefix_2018.xlsx"""
    result = []
    for fname in sorted(os.listdir(data_dir)):
        if fname.startswith("~"):
            continue
        if not fname.lower().endswith((".xls", ".xlsx")):
            continue
        m = re.search(r"_(\d{4})\.xlsx?$", fname, re.IGNORECASE)
        if m:
            result.append((int(m.group(1)), os.path.join(data_dir, fname)))
    return result


# ── Чтение Excel (один файл за раз) ───────────────────────────────────────────

def _read_xlsx_file(path: str) -> dict[str, dict]:
    """
    Читает .xlsx файл ОДИН РАЗ.
    Возвращает dict: sheet_name → {rows, merged_meta, range_address}.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    # Проход 1: данные (read_only=True — быстро, мало памяти)
    wb_data = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets_data: dict[str, list[list]] = {}
    for sheet_name in wb_data.sheetnames:
        rows = [list(row) for row in wb_data[sheet_name].iter_rows(values_only=True)]
        sheets_data[sheet_name] = rows
    wb_data.close()

    # Проход 2: merged cells (требует полной загрузки, но только один раз на файл)
    wb_full = openpyxl.load_workbook(path, data_only=True)
    merged: dict[str, str] = {}
    for sheet_name in wb_full.sheetnames:
        ws = wb_full[sheet_name]
        mc_list = []
        for mc in ws.merged_cells.ranges:
            cell_val = ws.cell(mc.min_row, mc.min_col).value
            mc_list.append({
                "range": str(mc),
                "value": str(cell_val).strip() if cell_val is not None else "",
            })
        merged[sheet_name] = json.dumps(mc_list, ensure_ascii=False)
    wb_full.close()

    # Сборка результата
    result = {}
    for sheet_name, rows in sheets_data.items():
        nrows = len(rows)
        ncols = max((len(r) for r in rows), default=0)
        col_letter = get_column_letter(ncols) if ncols > 0 else "A"
        result[sheet_name] = {
            "rows": rows,
            "merged_meta": merged.get(sheet_name, "[]"),
            "range_address": f"A1:{col_letter}{nrows}",
        }
    return result


def _read_xls_file(path: str) -> dict[str, dict]:
    """Читает .xls файл ОДИН РАЗ через xlrd."""
    import xlrd
    wb = xlrd.open_workbook(path, formatting_info=True)
    result = {}
    for sheet_name in wb.sheet_names():
        sh = wb.sheet_by_name(sheet_name)
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        mc_list = []
        for rlo, rhi, clo, chi in sh.merged_cells:
            val = sh.cell_value(rlo, clo)
            mc_list.append({
                "range": f"R{rlo}C{clo}:R{rhi - 1}C{chi - 1}",
                "value": str(val).strip() if val not in ("", None) else "",
            })
        result[sheet_name] = {
            "rows": rows,
            "merged_meta": json.dumps(mc_list, ensure_ascii=False),
            "range_address": f"R1C1:R{sh.nrows}C{sh.ncols}",
        }
    return result


def read_file(path: str) -> dict[str, dict]:
    """Читает Excel файл (xls или xlsx), возвращает данные всех листов."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return _read_xls_file(path)
    return _read_xlsx_file(path)


# ── Schema evolution ───────────────────────────────────────────────────────────

def ensure_columns(table, num_cols: int) -> None:
    existing = {f.name for f in table.schema().fields}
    missing = [i for i in range(num_cols) if f"col_{i}" not in existing]
    if not missing:
        return
    with table.update_schema() as upd:
        for i in missing:
            upd.add_column(f"col_{i}", StringType())
    log.info("Schema evolved: added col_%d..col_%d", missing[0], missing[-1])


# ── Запись батча листов в Iceberg ──────────────────────────────────────────────

def write_batch_to_bronze(
    catalog: SqlCatalog,
    load_id: str,
    loaded_at: datetime,
    source_file: str,
    year: int,
    sheets: dict[str, dict],   # sheet_name → {rows, merged_meta, range_address}
) -> dict[str, int]:
    """
    Пишет все листы из одного файла одним Iceberg append.
    Возвращает dict: sheet_name → количество записанных строк.
    """
    if not sheets:
        return {}

    # Максимальная ширина по всем листам (для schema evolution)
    max_cols = max(
        max((len(r) for r in s["rows"]), default=0)
        for s in sheets.values()
    )

    table = catalog.load_table("bronze.excel_tables")
    ensure_columns(table, max_cols)

    # Собираем все записи
    all_records: list[dict] = []
    counts: dict[str, int] = {}

    for sheet_name, sheet_data in sheets.items():
        rows = sheet_data["rows"]
        if not rows:
            counts[sheet_name] = 0
            continue
        ncols = max((len(r) for r in rows), default=0)
        for row_num, row in enumerate(rows):
            record: dict = {
                "load_id":           load_id,
                "loaded_at":         loaded_at,
                "source_file":       source_file,
                "year":              year,
                "sheet_name":        sheet_name,
                "row_num":           row_num,
                "range_address":     sheet_data["range_address"],
                "merged_cells_meta": sheet_data["merged_meta"],
            }
            for col_idx in range(ncols):
                val = row[col_idx] if col_idx < len(row) else None
                record[f"col_{col_idx}"] = str(val) if val is not None else None
            all_records.append(record)
        counts[sheet_name] = len(rows)

    if not all_records:
        return counts

    # Строим PyArrow таблицу
    current_schema = catalog.load_table("bronze.excel_tables").schema()
    pa_fields = []
    for field in current_schema.fields:
        nullable = not field.required
        if field.name == "loaded_at":
            pa_fields.append(pa.field(field.name, pa.timestamp("us", tz="UTC"), nullable=nullable))
        elif field.name in ("year", "row_num"):
            pa_fields.append(pa.field(field.name, pa.int32(), nullable=nullable))
        else:
            pa_fields.append(pa.field(field.name, pa.string(), nullable=nullable))

    pa_schema = pa.schema(pa_fields)
    col_names = [f.name for f in current_schema.fields]

    arrays = {
        col: pa.array([r.get(col) for r in all_records], type=pa_schema.field(col).type)
        for col in col_names
    }
    arrow_table = pa.table(arrays, schema=pa_schema)
    catalog.load_table("bronze.excel_tables").append(arrow_table)

    return counts


# ── Идемпотентность ────────────────────────────────────────────────────────────

def _load_existing_keys(catalog: SqlCatalog) -> set[tuple[str, str]]:
    """Одним сканом читает все уже загруженные (source_file, sheet_name)."""
    try:
        arrow = catalog.load_table("bronze.excel_tables").scan(
            selected_fields=("source_file", "sheet_name"),
        ).to_arrow()
        return set(zip(arrow["source_file"].to_pylist(), arrow["sheet_name"].to_pylist()))
    except Exception:
        return set()


# ── Точка входа ────────────────────────────────────────────────────────────────

def run(data_dir: str = DATA_DIR, flat: bool = False) -> dict:
    """
    Загружает все Excel-файлы из data_dir в Bronze.
    Идемпотентна: пропускает уже загруженные (source_file, sheet_name).

    flat=True: файлы прямо в папке, год из имени (_YYYY.xlsx).
    flat=False: файлы в подпапках с именем = год.
    """
    catalog = get_catalog()
    load_id = str(uuid.uuid4())
    loaded_at = datetime.now(tz=timezone.utc)
    stats: dict[tuple, int] = {}

    existing_keys = _load_existing_keys(catalog)
    log.info("Already loaded: %d (source_file, sheet_name) pairs", len(existing_keys))

    files = discover_files_flat(data_dir) if flat else discover_files(data_dir)

    for year, path in files:
        source_file = os.path.relpath(path, BASE_DIR).replace("\\", "/")
        log.info("Reading file: %s", source_file)

        # Читаем файл целиком (один раз)
        all_sheets = read_file(path)

        # Отсеиваем уже загруженные листы
        new_sheets = {
            name: data
            for name, data in all_sheets.items()
            if (source_file, name) not in existing_keys
        }

        skipped = len(all_sheets) - len(new_sheets)
        if skipped:
            log.info("  Skipping %d already loaded sheets", skipped)

        if not new_sheets:
            for name in all_sheets:
                stats[(year, os.path.basename(path), name)] = 0
            continue

        log.info("  Writing %d sheets as one batch...", len(new_sheets))
        counts = write_batch_to_bronze(
            catalog=catalog,
            load_id=load_id,
            loaded_at=loaded_at,
            source_file=source_file,
            year=year,
            sheets=new_sheets,
        )

        total_rows = sum(counts.values())
        log.info("  Done: %d rows written", total_rows)

        for name, count in counts.items():
            stats[(year, os.path.basename(path), name)] = count
        for name in all_sheets:
            if name not in counts:
                stats[(year, os.path.basename(path), name)] = 0

    return stats


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    result = run()
    print("\n── Итог ──────────────────────────")
    total = 0
    for (year, fname, sheet), count in sorted(result.items()):
        if count > 0:
            print(f"  {year}  {fname}  [{sheet}]  →  {count} строк")
        total += count
    print(f"  Всего новых строк: {total}")
