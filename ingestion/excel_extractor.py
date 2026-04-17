"""
Excel → pd.DataFrame extractor.
Только читает данные. Не пишет в БД.

Добавляет lineage-колонки к каждой строке:
  _etl_loaded_at — timestamp загрузки
  _source_file   — относительный путь к файлу
  _sheet_name    — имя листа
  _row_number    — порядковый номер строки в листе
  _year          — год (из имени папки или файла)
  _range_address — адрес диапазона листа
  _merged_cells_meta — JSON с описанием объединённых ячеек

Данные: col_0, col_1, ... col_N (имена фиксированы на этом уровне,
смысловые имена появляются в слое нормализации).
"""

import json
import os
import re
from datetime import datetime, timezone

import pandas as pd

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# ── Обнаружение файлов ─────────────────────────────────────────────────────────

def discover_files(data_dir: str) -> list[tuple[int, str]]:
    """Файлы в подпапках с именем = год: data_dir/2018/file.xlsx"""
    result = []
    for entry in sorted(os.scandir(data_dir), key=lambda e: e.name):
        if not entry.is_dir():
            continue
        try:
            year = int(entry.name)
        except ValueError:
            continue
        for fname in os.listdir(entry.path):
            if fname.lower().endswith((".xls", ".xlsx")) and not fname.startswith("~"):
                result.append((year, os.path.join(entry.path, fname)))
    return result


def discover_files_flat(data_dir: str) -> list[tuple[int, str]]:
    """Файлы прямо в папке, год из имени: data_dir/Prefix_2018.xlsx"""
    result = []
    for fname in sorted(os.listdir(data_dir)):
        if fname.startswith("~"):
            continue
        if not fname.lower().endswith((".xls", ".xlsx")):
            continue
        m = re.search(r"_(\d{4})\.xlsx?$", fname, re.IGNORECASE)
        if m:
            result.append((int(m.group(1)), os.path.join(data_dir, fname)))
    return result


# ── Чтение Excel ───────────────────────────────────────────────────────────────

def _read_xlsx_file(path: str) -> dict[str, dict]:
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb_data = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets_data: dict[str, list[list]] = {}
    for sheet_name in wb_data.sheetnames:
        rows = [list(row) for row in wb_data[sheet_name].iter_rows(values_only=True)]
        sheets_data[sheet_name] = rows
    wb_data.close()

    wb_full = openpyxl.load_workbook(path, data_only=True)
    merged: dict[str, str] = {}
    for sheet_name in wb_full.sheetnames:
        ws = wb_full[sheet_name]
        mc_list = []
        for mc in ws.merged_cells.ranges:
            cell_val = ws.cell(mc.min_row, mc.min_col).value
            mc_list.append({
                "range": str(mc),
                "value": str(cell_val).strip() if cell_val is not None else "",
            })
        merged[sheet_name] = json.dumps(mc_list, ensure_ascii=False)
    wb_full.close()

    result = {}
    for sheet_name, rows in sheets_data.items():
        nrows = len(rows)
        ncols = max((len(r) for r in rows), default=0)
        col_letter = get_column_letter(ncols) if ncols > 0 else "A"
        result[sheet_name] = {
            "rows": rows,
            "merged_meta": merged.get(sheet_name, "[]"),
            "range_address": f"A1:{col_letter}{nrows}",
        }
    return result


def _read_xls_file(path: str) -> dict[str, dict]:
    import xlrd
    wb = xlrd.open_workbook(path, formatting_info=True)
    result = {}
    for sheet_name in wb.sheet_names():
        sh = wb.sheet_by_name(sheet_name)
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        mc_list = []
        for rlo, rhi, clo, chi in sh.merged_cells:
            val = sh.cell_value(rlo, clo)
            mc_list.append({
                "range": f"R{rlo}C{clo}:R{rhi - 1}C{chi - 1}",
                "value": str(val).strip() if val not in ("", None) else "",
            })
        result[sheet_name] = {
            "rows": rows,
            "merged_meta": json.dumps(mc_list, ensure_ascii=False),
            "range_address": f"R1C1:R{sh.nrows}C{sh.ncols}",
        }
    return result


def _read_file(path: str) -> dict[str, dict]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return _read_xls_file(path)
    return _read_xlsx_file(path)


# ── Точка входа ────────────────────────────────────────────────────────────────

def extract(data_dir: str, flat: bool = False) -> pd.DataFrame:
    """
    Читает все Excel-файлы из data_dir.
    Возвращает pd.DataFrame со всеми строками и lineage-колонками.

    flat=True: файлы прямо в папке, год из имени (_YYYY.xlsx).
    flat=False: файлы в подпапках с именем = год.
    """
    loaded_at = datetime.now(tz=timezone.utc)
    files = discover_files_flat(data_dir) if flat else discover_files(data_dir)

    all_rows: list[dict] = []
    for year, path in files:
        source_file = os.path.relpath(path, BASE_DIR).replace("\\", "/")
        all_sheets = _read_file(path)

        for sheet_name, sheet_data in all_sheets.items():
            rows = sheet_data["rows"]
            for row_number, row in enumerate(rows):
                record: dict = {
                    "_etl_loaded_at":    loaded_at,
                    "_source_file":      source_file,
                    "_sheet_name":       sheet_name,
                    "_row_number":       row_number,
                    "_year":             year,
                    "_range_address":    sheet_data["range_address"],
                    "_merged_cells_meta": sheet_data["merged_meta"],
                }
                for col_idx, val in enumerate(row):
                    record[f"col_{col_idx}"] = str(val) if val is not None else None
                all_rows.append(record)

    if not all_rows:
        return pd.DataFrame()

    return pd.DataFrame(all_rows)
